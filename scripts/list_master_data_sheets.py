import re
from openpyxl import load_workbook

wb = load_workbook('/home/ubuntu/upload/DCT_Master_Data_Intake.xlsx', read_only=True, data_only=False)
for position, ws in enumerate(wb.worksheets, 1):
    name = ws.title
    lowered = name.lower()
    if re.search(r'(^|[^a-z])old([^a-z]|$)', lowered):
        status = 'HISTORICAL'
    elif 'on hold' in lowered or 'not generated' in lowered:
        status = 'NON-CURRENT'
    elif name in {'README', 'Master Data Authoring Guide', 'Load Order', 'Conventions', 'Sheet1'}:
        status = 'GOVERNANCE/REFERENCE'
    else:
        status = 'ACTIVE'
    print(f'{position:02d}|{status}|{name}|{ws.max_row} rows|{ws.max_column} cols')
