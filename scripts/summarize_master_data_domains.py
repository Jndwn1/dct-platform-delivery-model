from collections import Counter
from openpyxl import load_workbook

WORKBOOK = '/home/ubuntu/upload/DCT_Master_Data_Intake.xlsx'
wb = load_workbook(WORKBOOK, data_only=False, read_only=False)

def text(value):
    return '' if value is None else str(value).strip()

def fill(cell):
    color = cell.fill.fgColor
    if color.type == 'rgb' and color.rgb and color.rgb != '00000000':
        return color.rgb[-6:]
    if color.type == 'indexed' and color.indexed not in (None, 64):
        return f'indexed:{color.indexed}'
    return None

for ws in wb.worksheets:
    candidate = (0, [])
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        values = [text(v) for v in row]
        score = sum(1 for v in values if v)
        if score > candidate[0]:
            candidate = (score, values)
    header_row = next((idx for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True), 1) if [text(v) for v in row] == candidate[1]), 1)
    headers = [v for v in candidate[1] if v][:12]
    fills = Counter()
    for row in ws.iter_rows():
        for cell in row:
            name = fill(cell)
            if name:
                fills[name] += 1
    classification = 'HISTORICAL' if 'old' in ws.title.lower() else 'ACTIVE'
    print(f"{classification} | {ws.title} | rows={ws.max_row} | header_row={header_row} | headers={' ; '.join(headers)} | fills={dict(fills.most_common(6))}")

for target in ('TDC - Adjustment Rules', 'TDC - Adjustment Rule Inputs', 'TDC - Adjustment Rule Lines'):
    ws = wb[target]
    print(f"\nDETAIL | {target}")
    for row in ws.iter_rows():
        values = [text(c.value) for c in row]
        if any(code in ' | '.join(values) for code in ('MP-02', 'MP-06', 'MP-07', 'MP-08')):
            colored = [f"{c.coordinate}:{text(c.value)}:{fill(c) or 'none'}" for c in row if text(c.value)]
            print(' || '.join(colored[:18]))
