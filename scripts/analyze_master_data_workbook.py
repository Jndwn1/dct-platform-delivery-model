from collections import Counter
from datetime import date, datetime
import json
from pathlib import Path
from openpyxl import load_workbook

WORKBOOK = Path('/home/ubuntu/upload/DCT_Master_Data_Intake.xlsx')

def clean(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None:
        return None
    return str(value).strip()

def fill_name(cell):
    color = cell.fill.fgColor
    if color.type == 'rgb' and color.rgb and color.rgb != '00000000':
        return color.rgb[-6:]
    if color.type == 'indexed' and color.indexed not in (None, 64):
        return f'indexed:{color.indexed}'
    if color.type == 'theme' and color.theme is not None:
        return f'theme:{color.theme}'
    return None

wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
summary = []
for ws in wb.worksheets:
    nonempty_rows = []
    for row in ws.iter_rows(values_only=False):
        values = [clean(c.value) for c in row]
        if any(v not in (None, '') for v in values):
            nonempty_rows.append((row, values))
    headers = nonempty_rows[0][1] if nonempty_rows else []
    headers = [h for h in headers if h not in (None, '')]
    preview = []
    for row, values in nonempty_rows[1:4]:
        preview.append([v for v in values[:12]])
    fills = Counter()
    colored_examples = []
    for row, values in nonempty_rows:
        for cell in row:
            name = fill_name(cell)
            if name:
                fills[name] += 1
                if len(colored_examples) < 5 and clean(cell.value):
                    colored_examples.append({'cell': cell.coordinate, 'value': clean(cell.value), 'fill': name})
    summary.append({
        'sheet': ws.title,
        'classification': 'historical' if 'old' in ws.title.lower() else 'active',
        'dimensions': {'rows': ws.max_row, 'columns': ws.max_column, 'nonempty_rows': len(nonempty_rows)},
        'headers': headers[:20],
        'preview_rows': preview,
        'fill_counts': dict(fills.most_common()),
        'colored_examples': colored_examples,
    })

print(json.dumps({'workbook': WORKBOOK.name, 'sheets': summary}, indent=2))
